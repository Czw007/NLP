#收集数据——去除数据中非文本部分（html标签等）——分词——去除停用词Stopping words——词干提取stemming和词型还原——英文单词转换大小写——特征提取td-idf
from openpyxl import load_workbook
import string

#读取数据
def readData():
    ws=load_workbook("configuration_stackoverflow.xlsx")
    wa=ws.active
    max_row=wa.max_row
    data=[]
    for i in range(2,max_row+1):
        data.append(str(wa.cell(i,7).value))
    return data
    # print(len(data))#len=2575
    # print(data)
def writeDate():
    ws = load_workbook("configuration_stackoverflow.xlsx")
    wa = ws.active
    max_row = wa.max_row
#去除标点符号
def punctuation(text_list):
    res=[]
    punctuation_string=string.punctuation
    for text in text_list:
        for i in punctuation_string:
            text=text.replace(i,'')
        res.append(text)
    return res


#分词，英文分词直接split行数，中文使用jieba库
def en_tokensize(text):
    return text.split(' ')
#去除停用词
from nltk.corpus import stopwords
def remover_stopWords(text_list):
    res=[]
    for text in text_list:
        filter_text = [w for w in text.split(' ') if w not in stopwords.words('english')]
        res.append(' '.join(filter_text))
    return res

#词干提取和词型还原stemming and lemmanzation
from nltk.stem import WordNetLemmatizer
from nltk.stem import SnowballStemmer
#词干提取是去除单词的前后缀得到词根的过程,词形还原是基于词典，将单词的复杂形态转变成最基础的形态
#词干提取
def stemming(text_list):
    stemmer = SnowballStemmer("english")  # 选择语言
    res = []
    for text in text_list:
        text_t=[]
        for word in text.split(' '):
            text_t.append(stemmer.stem(word))
        res.append(' '.join(text_t))
    return res
#词形还原
def lemmanzation(text_list):
    wnl = WordNetLemmatizer()
    res = []
    for text in text_list:
        text_t=[]
        for word in text.split(' '):
            text_t.append(wnl.lemmatize(word))
        res.append(' '.join(text_t))
    return res
#字母转换成小写
def toLower(text_list):
    res = []
    for text in text_list:
        text_t=[]
        for word in text.split(' '):
            text_t.append(word.lower())
        res.append(' '.join(text_t))
    return res

#读取数据
import datetime
startTime=datetime.datetime.now()
data=readData()
print("data:\n",data[0])
#去除停用词（会改变语义
data_remover_stopwords=remover_stopWords(data)
print("data_remover_stopwords:\n",data_remover_stopwords[0])
#词干提取和词型还原stemming and lemmanzation
data_lemmanzation=lemmanzation(data_remover_stopwords)
print("data_lemmanzation:\n",data_lemmanzation[0])
#字母转换成大小写
data_tolower=toLower(data_lemmanzation)
print("data_tolower:\n",data_tolower[0])

#TF-IDF基于词频将文本转换成向量
from sklearn.feature_extraction.text import TfidfVectorizer
import pandas as pd
vector = TfidfVectorizer()
tf_data = vector.fit_transform(data_tolower)
print("tf_data:\n",tf_data)#(句子下标, 单词特征下标)   权重
print(vector.vocabulary_)    #单词特征

df1 = pd.DataFrame(tf_data.toarray(), columns=vector.get_feature_names()) # to DataFrame
print(df1)
#聚类Kmeans
from sklearn.cluster import KMeans

num_clusters = 2
km_cluster = KMeans(n_clusters=num_clusters, max_iter=300, n_init=1, \
                    init='k-means++',n_jobs=1)
result = km_cluster.fit_predict(tf_data)
print("result:\n",result)
for i,res in enumerate(result):
    print(i,":",res)
endTime=datetime.datetime.now()
print("运行总时间是：\n",(endTime-startTime).seconds,"秒")
