from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl import Workbook
# 创建一个Workbook
wb = Workbook() #默认生成一个名为'Sheet'的WorkSheet
# 获取默认打开的(active)的WorkSheet
wb.save('stackoverflow_configuration.xlsx')

mybook = load_workbook('stackoverflow_configuration.xlsx')
sheet_names = mybook.sheetnames
print(sheet_names)
ws=mybook.create_sheet("Sheet{}".format(2))
ws.cell(1,1).value="Sheet2"
print(ws.cell(1,1).value)

ws=mybook.create_sheet("Sheet{}".format(3))
ws.cell(1,1).value="Sheet3"
print(ws.cell(1,1).value)

sheet_names = mybook.sheetnames
print(sheet_names)
mybook.save('stackoverflow_configuration.xlsx')
print("over")

# ws.save('stackoverflow_configuration.xlsx')